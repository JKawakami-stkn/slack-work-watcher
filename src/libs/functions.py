# -*- coding: utf-8 -*-
# 日時
import datetime
import time
# mysql
import mysql.connector

# excel操作
import openpyxl as excel
# from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles import Border, Side, PatternFill




###########################################################################
### 名前      ：isQueryForRead
### 説明      ：セレクト文かどうか
### 引数      ：(str)クエリ
### 戻り値    ：(bool)結果
### 呼ぶ関数  ：
###########################################################################
def isQueryForRead(query):

    if(query.startswith("SELECT") or query.startswith("select")):
        return True
    else:
        return False



###########################################################################
### 名前      ：connectDatabase
### 説明      ：DBに接続する
### 引数      ：なし
### 戻り値    ：接続情報インスタンス
### 呼ぶ関数  ：
###########################################################################
def connectDatabase():
    # 接続先サーバーを指定
    db=mysql.connector.connect(host="localhost", user="root", password="password" ,charset='utf8')
    cursor=db.cursor(buffered=True)

    # 接続先dbを指定
    cursor.execute("USE job_watcher")
    # 接続
    db.commit()

    return db, cursor



###########################################################################
### 名前      ：executionQuery
### 説明      ：クエリを実行する
### 引数      ：(str)実行するクエリ
### 戻り値    ：エラーでNone、成功でデータ
### 呼ぶ関数  ：db_commit(), isSelectSyntax(str)
###########################################################################
def executionQuery(query):

    result = True

    try:

        db, cursor = connectDatabase()

        cursor.execute(query)

        if(isQueryForRead(query)):
            result = cursor.fetchall()

        db.commit()

        return result

    except mysql.connector.errors:

        return False

    finally:
        cursor.close()
        db.close()



###########################################################################
### 名前      ：trimQueryResult
### 説明      ：クエリ結果を整形する
### 引数      ：{}
### 戻り値    ：整形結果
### 参照関数  ：
###########################################################################
def trimQueryResult(query_result, column_name=None):


    if(type(query_result) == "bool"):
        return query_result

    if(len(query_result) == 1 and len(query_result[0]) == 1):
        return query_result[0][0]

    if(column_name is None):
        return query_result

    trim_data = []

    for record in query_result:

        trim_task_data = {}

        for i, field in enumerate(record):

            trim_task_data[column_name[i]] = field

        trim_data.append(trim_task_data)


    return trim_data






###########################################################################
### 名前      ：createWBS
### 説明      ：CSVファイルを作成
### 引数      ：(user_id)取得するユーザーのID
### 戻り値    ：終了予定日
### 参照関数  ：getUserName()
###########################################################################

# TODO : 予定より開始が早いタスクと予定より遅れているタスクをぱっと見で分かるように色分けする
def createWBS():

    # 新規ワークブックを作成
    wb = excel.Workbook()
    ws = wb.active

    # column = ["No", "タスク", "", "開始予定日", "終了予定日", "開始日", "終了日", "担当者", "ステータス"]
    column = ["No", "タスク", "担当者", "開始日", "終了日", "開始予定日", "終了予定日",  "ステータス"]

    # セルの枠線
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 操作対処のセル
    this_cell = None

    # プロジェクト開始日
    project_start_date = None

    # ヘッダー部を作成
    for i in range(1, len(column) + 1):

        # セルを取得
        this_cell = ws.cell(column = i, row = 2, value = column[i - 1])
        this_cell.fill = PatternFill(fill_type='solid',fgColor='ffde8c') # セルの色
        this_cell.border = border # 枠線

        # 下のセルと結合
        ws.merge_cells( start_row = 2, start_column = i, end_row = 3, end_column=i )

    # タスク情報を取得
    query_to_get_all_user = "SELECT * FROM task;"
    query_to_get_all_user_result = executionQuery( query_to_get_all_user )
    tasks = query_to_get_all_user_result

    # キー:ユーザーID, 値:ユーザ名 の辞書を作成
    users_data_dict = {}
    query_to_get_all_user = "SELECT * FROM user;"
    query_to_get_all_user_result = executionQuery( query_to_get_all_user )
    users_data = query_to_get_all_user_result

    for user_data in users_data:
        users_data_dict.setdefault(user_data[0], user_data[1])

    # タスク情報書き込み
    for i, task in enumerate(tasks): # 縦
        for j, data in enumerate(task): # 横

            # TODO : 変わるvalueのみを変数に格納しborder処理を減らす

            width =  i + 4
            height = j + 1

            # プロジェクトの開始日を取得
            if(i == 0 and j== 3):
                project_start_date = data

            # ユーザー名
            if(j == 2):
                # ユーザー名を表示する処理
                if(data in users_data_dict):
                    this_cell = ws.cell(column = height, row = width, value = users_data_dict[data])
                    this_cell.border = border # 枠線
                    continue

            # 終了日
            if(j == 4):
                # ステータス設定
                if(data is not None):
                    this_cell = ws.cell(column = height, row = width, value = str(data))
                    this_cell.border = border # 枠線
                    this_cell = ws.cell(column = len(task) + 1, row = width, value = "完了")
                    this_cell.border = border # 枠線
                    continue
                else:
                    this_cell = ws.cell(column = height, row = width, value = "")
                    this_cell.border = border # 枠線
                    this_cell = ws.cell(column = len(task) + 1, row = width, value = "作業中")
                    this_cell.border = border # 枠線
                    continue

            this_cell = ws.cell(column = j + 1, row = width, value = str(data))

            this_cell.border = border # 枠線

    # 曜日名
    weekname = ["月","火","水","木","金","土", "日"]

    # 月
    this_month = 0



    # 366日分を繰り返してセルに書き込む
    tm = project_start_date

    # 月の始まりのセル
    start_month_width  = 9

    for i in range(1, 367):

        width =  i + 8
        height = 1

        # 月初めのみ出力
        if(this_month != tm.month):
            this_cell = ws.cell(column = width, row = height, value=tm.month)
            this_month = tm.month
            this_cell.border = border # 枠線
            # 最初の月じゃない場合に結合
            # FIXME : 最後の月が結合されない
            if(i != 1):
                # セルと結合
                ws.merge_cells( start_row = height, start_column = start_month_width, end_row = height, end_column = width - 1)
                start_month_width = width

        this_cell = ws.cell(column=i + 8, row=2, value=tm.day)
        this_cell.border = border # 枠線
        this_cell = ws.cell(column=i + 8, row=3, value=weekname[tm.weekday()])
        this_cell.border = border # 枠線

        # 翌日
        tm = tm + datetime.timedelta(days=1)


    # プロジェクト開始日のセル
    project_start_sell = 9
    project_start_date

    # 予定出力
    for i, data in enumerate(tasks):

        # タスク開始予定セルを取得
        start_cell = (data[5] - project_start_date).days + project_start_sell

        # タスク終了予定セルを取得
        finish_cell = (data[6] - project_start_date).days + project_start_sell

        # せる塗りつぶし
        for j in range(start_cell, finish_cell + 1):
            this_cell = ws.cell(column = j, row = i + 4)
            this_cell.fill = PatternFill(fill_type='solid',fgColor='f4ff78') # セルの色


    # 実績出力
    for i, data in enumerate(tasks):

        # タスク開始セルを取得
        start_cell = (data[3] - project_start_date).days + project_start_sell

        # タスク終了セルを取得
        if(data[4] is not None):
            finish_cell = (data[4] - project_start_date).days + project_start_sell
        else:
            finish_cell = (datetime.date.today() - project_start_date).days + project_start_sell

        # せる塗りつぶし
        for j in range(start_cell, finish_cell + 1):
            this_cell = ws.cell(column = j, row = i + 4)
            this_cell.fill = PatternFill(fill_type='solid',fgColor='4aaeff') # セルの色


    # 保存
    now = time.time()
    path = "../files/WBS" + str(now) + ".xlsx"
    wb.save(path)


    return(path)






